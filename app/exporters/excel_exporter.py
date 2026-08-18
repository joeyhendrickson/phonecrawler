from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.exporters.csv_exporter import (
    errors_to_rows,
    inventory_to_rows,
    occurrences_to_rows,
    url_inventory_to_rows,
)
from app.exporters.summary import summary_rows, top_phones_by_occurrences, top_phones_by_urls
from app.models.records import CrawlResult

PHONE_COLUMNS = {
    "raw_phone",
    "normalized_phone",
    "e164_phone",
    "national_format",
    "extension",
}
URL_COLUMNS = {
    "source_url",
    "final_url",
    "first_seen_url",
    "requested_url",
    "url",
    "referring_url",
    "canonical_url",
    "sitemap_url",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BAND_FILL = PatternFill("solid", fgColor="D6E3F0")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")


def export_excel(result: CrawlResult, output_dir: Path) -> Path:
    import pandas as pd

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "phone_inventory.xlsx"
    coverage = url_inventory_to_rows(result.url_inventory)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows(result)).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(inventory_to_rows(result.unique_phones)).to_excel(
            writer, sheet_name="Phone Inventory", index=False
        )
        pd.DataFrame(occurrences_to_rows(result.occurrences)).to_excel(
            writer, sheet_name="Phone Occurrences", index=False
        )
        pd.DataFrame(coverage).to_excel(writer, sheet_name="URL Inventory", index=False)
        pd.DataFrame(coverage).to_excel(writer, sheet_name="Crawl Coverage", index=False)
        pd.DataFrame(errors_to_rows(result)).to_excel(writer, sheet_name="Errors", index=False)

        workbook = writer.book
        _append_top_phone_tables(
            workbook["Summary"],
            top_phones_by_occurrences(result.unique_phones),
            top_phones_by_urls(result.unique_phones),
        )
        for name in (
            "Summary",
            "Phone Inventory",
            "Phone Occurrences",
            "URL Inventory",
            "Crawl Coverage",
            "Errors",
        ):
            _style_sheet(workbook[name], enable_filter=(name != "Summary"))
    return path


def _append_top_phone_tables(sheet, frequent: list[dict], most_urls: list[dict]) -> None:
    start = sheet.max_row + 2
    sheet.cell(start, 1, "Most frequently published phone numbers").font = TITLE_FONT
    _write_table(sheet, frequent, start_row=start + 1)
    start = sheet.max_row + 2
    sheet.cell(start, 1, "Phone numbers appearing on the most URLs").font = TITLE_FONT
    _write_table(sheet, most_urls, start_row=start + 1)


def _write_table(sheet, rows: list[dict], *, start_row: int) -> None:
    if not rows:
        sheet.cell(start_row, 1, "(none)")
        return
    headers = list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, col, header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = sheet.cell(row_index, col, value)
            if header in PHONE_COLUMNS and value is not None:
                cell.number_format = "@"
                cell.value = str(value)


def _style_sheet(sheet, *, enable_filter: bool) -> None:
    if sheet.max_row == 0 or sheet.max_column == 0:
        return
    sheet.freeze_panes = "A2"
    if enable_filter:
        last_col = get_column_letter(max(sheet.max_column, 1))
        # Filter the header row of the primary table only.
        sheet.auto_filter.ref = f"A1:{last_col}{max(sheet.max_row, 1)}"

    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
        header_row = sheet[1]
        for cell in row:
            header = header_row[cell.column - 1].value
            if header in PHONE_COLUMNS and cell.value is not None:
                cell.number_format = "@"
                cell.value = str(cell.value)
            if header in URL_COLUMNS and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")

    for index in range(1, sheet.max_column + 1):
        letter = get_column_letter(index)
        max_len = 0
        for cell in sheet[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 55))
        sheet.column_dimensions[letter].width = max(14, max_len + 2)
    sheet.row_dimensions[1].height = 22
